#!/usr/bin/env python3
"""Render an .xlsx workbook summarizing a VCF.

Sheets: summary, by_chromosome, top_quality, frameshift_candidates, filtered_out.
Output written next to the input as <input>.variants.xlsx (path printed to stdout).
"""

from __future__ import annotations

import sys
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _common import detect_format, die  # noqa: E402

from openpyxl import Workbook  # noqa: E402


def _is_frameshift(ref: str, alt: str) -> bool:
    if not ref or not alt:
        return False
    diff = abs(len(ref) - len(alt))
    return diff > 0 and diff % 3 != 0


def main(argv: list[str]) -> int:
    if len(argv) != 2 or argv[1] in {"-h", "--help"}:
        print("usage: variants_report.py <vcf>", file=sys.stderr)
        return 2
    path = argv[1]
    try:
        info = detect_format(path)
    except FileNotFoundError:
        die(f"no such file: {path}")
    if info.format != "vcf":
        die(f"expected VCF, got {info.format}")

    import cyvcf2
    vcf = cyvcf2.VCF(path)

    summary_counts = Counter()
    per_chrom = Counter()
    top_quality: list[tuple[float, str, int, str, str, str]] = []  # (qual, chrom, pos, id, ref, alt)
    frameshifts: list[tuple[str, int, str, str, str]] = []
    filtered: list[tuple[str, int, str, str, str, str]] = []

    for v in vcf:
        per_chrom[v.CHROM] += 1
        is_pass = v.FILTER is None or v.FILTER == "PASS"
        if is_pass:
            summary_counts["pass"] += 1
        else:
            summary_counts["filtered"] += 1
        ref = v.REF or ""
        alt = (v.ALT or [""])[0]
        if v.INFO.get("SVTYPE"):
            summary_counts["sv"] += 1
        elif len(ref) == 1 and len(alt) == 1:
            summary_counts["snv"] += 1
        elif len(ref) == len(alt):
            summary_counts["mnp"] += 1
        else:
            summary_counts["indel"] += 1
        q = v.QUAL if v.QUAL is not None else 0.0
        top_quality.append((q, v.CHROM, v.POS, v.ID or ".", ref, alt))
        if _is_frameshift(ref, alt):
            frameshifts.append((v.CHROM, v.POS, v.ID or ".", ref, alt))
        if not is_pass:
            filtered.append((v.CHROM, v.POS, v.ID or ".", ref, alt, v.FILTER or ""))

    top_quality.sort(reverse=True)
    top_quality = top_quality[:100]

    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    ws.append(["metric", "value"])
    for k in ("snv", "indel", "mnp", "sv", "pass", "filtered"):
        ws.append([k, summary_counts.get(k, 0)])
    ws.append(["samples", ",".join(vcf.samples)])

    ws2 = wb.create_sheet("by_chromosome")
    ws2.append(["chrom", "count"])
    for c, n in sorted(per_chrom.items()):
        ws2.append([c, n])

    ws3 = wb.create_sheet("top_quality")
    ws3.append(["qual", "chrom", "pos", "id", "ref", "alt"])
    for row in top_quality:
        ws3.append(list(row))

    ws4 = wb.create_sheet("frameshift_candidates")
    ws4.append(["chrom", "pos", "id", "ref", "alt"])
    for row in frameshifts:
        ws4.append(list(row))

    ws5 = wb.create_sheet("filtered_out")
    ws5.append(["chrom", "pos", "id", "ref", "alt", "filter"])
    for row in filtered:
        ws5.append(list(row))

    out_path = Path(path).with_suffix(Path(path).suffix + ".variants.xlsx")
    if out_path.name.endswith(".gz.variants.xlsx"):
        out_path = Path(str(out_path).replace(".gz.variants.xlsx", ".variants.xlsx"))
    wb.save(out_path)
    print(str(out_path))
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
